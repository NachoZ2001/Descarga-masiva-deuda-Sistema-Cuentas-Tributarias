from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, NamedStyle, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd
import time
import pyautogui
import os
import glob
import xlwings as xw

# Definir rutas a las carpetas y archivos
input_folder_excel = "C:/Program Files/Sublime Merge/Descarga-masiva-deuda-Sistema-de-Cuentas-Tributarias/data/input/Deudas"
output_folder_csv = "C:/Program Files/Sublime Merge/Descarga-masiva-deuda-Sistema-de-Cuentas-Tributarias/data/input/DeudasCSV"
output_file_csv = "C:/Program Files/Sublime Merge/Descarga-masiva-deuda-Sistema-de-Cuentas-Tributarias/data/Resumen_deudas.csv"
output_file_xlsx = "C:/Program Files/Sublime Merge/Descarga-masiva-deuda-Sistema-de-Cuentas-Tributarias/data/Resumen_deudas.xlsx"
fecha_especifica = '2024-03-31'

# Leer el archivo Excel
df = pd.read_excel(r'C:\Program Files\Sublime Merge\Descarga-masiva-deuda-Sistema-de-Cuentas-Tributarias\data\input\clientes.xlsx')

# Suposición de nombres de columnas
cuit_login_list = df['CUIT para ingresar'].tolist()
cuit_represent_list = df['CUIT representado'].tolist()
password_list = df['Contraseña'].tolist()
download_list = df['Ubicacion Descarga'].tolist()
posterior_list = df['Posterior'].tolist()
anterior_list = df['Anterior'].tolist()
clientes_list = df['Cliente'].tolist()

# Configuración de opciones de Chrome
options = Options()
options.add_argument("--start-maximized")
options.add_argument("--disable-blink-features=AutomationControlled")
options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36")
options.add_experimental_option("excludeSwitches", ["enable-automation"])
options.add_experimental_option('useAutomationExtension', False)

# Configurar preferencias de descarga
prefs = {
    "download.prompt_for_download": True,
    "download.directory_upgrade": True,
    "safebrowsing.enabled": True
}
options.add_experimental_option("prefs", prefs)

# Inicializar driver
service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=service, options=options)

def iniciar_sesion(cuit_ingresar, password):
    """Inicia sesión en el sitio web con el CUIT y contraseña proporcionados."""
    try:
        driver.get('https://auth.afip.gob.ar/contribuyente_/login.xhtml')
        element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, 'F1:username')))
        element.clear()
        time.sleep(5)

        element.send_keys(cuit_ingresar)
        driver.find_element(By.ID, 'F1:btnSiguiente').click()
        time.sleep(5)

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, 'F1:password'))).send_keys(password)
        time.sleep(5)
        driver.find_element(By.ID, 'F1:btnIngresar').click()
        time.sleep(5)
    except Exception as e:
        print(f"Error al iniciar sesión: {e}")

def ingresar_modulo():
    """Ingresa al módulo específico del sistema de cuentas tributarias."""
    try:
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.LINK_TEXT, "Ver todos"))).click()
        time.sleep(5)
        driver.find_element(By.ID, 'buscadorInput').send_keys('SISTEMA DE CUENTAS TRIBUTARIAS')
        time.sleep(5)
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, 'rbt-menu-item-0'))).click()
        time.sleep(10)

        # Cambiar de pestaña
        window_handles = driver.window_handles
        driver.switch_to.window(window_handles[-1])

        try:
            error_message = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.TAG_NAME, 'pre')))
            if error_message.text == "Ha ocurrido un error al autenticar, intente nuevamente.":
                driver.refresh()
                time.sleep(5)
        except:
            pass
    except Exception as e:
        print(f"Error al ingresar al módulo: {e}")

def seleccionar_cuit_representado(cuit_representado):
    """Selecciona el CUIT representado en el sistema."""
    try:
        select_present = EC.presence_of_element_located((By.NAME, "$PropertySelection"))
        if WebDriverWait(driver, 5).until(select_present):
            current_selection = Select(driver.find_element(By.NAME, "$PropertySelection")).first_selected_option.text
            if current_selection != str(cuit_representado):
                select_element = Select(driver.find_element(By.NAME, "$PropertySelection"))
                select_element.select_by_visible_text(str(cuit_representado))
    except Exception:
        try:
            cuit_element = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'span.cuit')))
            cuit_text = cuit_element.text.replace('-', '')
            if cuit_text != str(cuit_representado):
                print(f"El CUIT ingresado no coincide con el CUIT representado: {cuit_representado}")
                return False
        except Exception as e:
            print(f"Error al verificar CUIT: {e}")
            return False
    return True

def exportar_excel(ubicacion_descarga, cuit_representado, cliente, cantidad_faltas_presentacion):
    """Descarga y guarda el archivo Excel en la ubicación especificada."""
    try:       
        # Exportar XLSX
        WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, "//*[@id='DataTables_Table_0_wrapper']/div[1]/a[2]/span"))).click()
        time.sleep(5)

        # Guardarlo con nombre y carpeta especifica
        
        nombre_archivo = f"{cuit_representado}- {cliente} - {cantidad_faltas_presentacion} - Deudas"
        pyautogui.write(nombre_archivo)
        time.sleep(1)
        pyautogui.hotkey('alt', 'd')
        time.sleep(0.5)
        pyautogui.write(ubicacion_descarga)
        time.sleep(1)
        pyautogui.press('enter')
        time.sleep(1)
        pyautogui.hotkey('alt', 't')
        time.sleep(1)
        pyautogui.press('enter')
        time.sleep(1)
    except Exception as e:
        print(f"Error al exportar el archivo Excel: {e}")

def cerrar_sesion():
    """Cierra la sesión actual."""
    try:
        driver.close()
        window_handles = driver.window_handles
        driver.switch_to.window(window_handles[0])
        driver.find_element(By.ID, "iconoChicoContribuyenteAFIP").click()
        driver.find_element(By.XPATH, '//*[@id="contBtnContribuyente"]/div[6]/button/div/div[2]').click()
        time.sleep(5)
    except Exception as e:
        print(f"Error al cerrar sesión: {e}")

def extraer_datos_nuevo(cuit_ingresar, cuit_representado, password, ubicacion_descarga, posterior, cliente):
    """Extrae datos para un nuevo usuario."""
    try:
        iniciar_sesion(cuit_ingresar, password)
        ingresar_modulo()
        if seleccionar_cuit_representado(cuit_representado):
            cantidad_faltas_presentacion = driver.find_element(By.NAME, "functor$1").get_attribute('value')
            exportar_excel(ubicacion_descarga, cuit_representado, cliente, cantidad_faltas_presentacion)
            if posterior == 0:
                cerrar_sesion()
            return cantidad_faltas_presentacion
    except Exception as e:
        print(f"Error al extraer datos para el nuevo usuario: {e}")

def extraer_datos(cuit_representado, ubicacion_descarga, posterior, cliente):
    """Extrae datos para un usuario existente."""
    try:
        if seleccionar_cuit_representado(cuit_representado):
            cantidad_faltas_presentacion = driver.find_element(By.NAME, "functor$1").get_attribute('value')
            exportar_excel(ubicacion_descarga, cuit_representado, cliente, cantidad_faltas_presentacion)
            if posterior == 0:
                cerrar_sesion()
    except Exception as e:
        print(f"Error al extraer datos: {e}")

# Crear el archivo de resultados
resultados = []

# Iterar sobre cada cliente
for cuit_ingresar, cuit_representado, password, download, posterior, anterior, cliente in zip(cuit_login_list, cuit_represent_list, password_list, download_list, posterior_list, anterior_list, clientes_list):
    if anterior == 0:
        extraer_datos_nuevo(cuit_ingresar, cuit_representado, password, download, posterior, cliente)
    else:
        extraer_datos(cuit_representado, download, posterior, cliente)

# Crear la carpeta de salida para CSV si no existe
os.makedirs(output_folder_csv, exist_ok=True)

# Función para convertir Excel a CSV utilizando xlwings
def excel_a_csv(input_folder, output_folder):
    for excel_file in glob.glob(os.path.join(input_folder, "*.xlsx")):
        try:
            app = xw.App(visible=False)
            wb = app.books.open(excel_file)
            sheet = wb.sheets[0]
            df = sheet.used_range.options(pd.DataFrame, header=1, index=False).value

            # Convertir la columna 'FechaVencimiento' a datetime, ajustar según sea necesario
            if 'FechaVencimiento' in df.columns:
                df['FechaVencimiento'] = pd.to_datetime(df['FechaVencimiento'], errors='coerce')

            wb.close()
            app.quit()

            base = os.path.basename(excel_file)
            csv_file = os.path.join(output_folder, base.replace('.xlsx', '.csv'))
            df.to_csv(csv_file, index=False, encoding='utf-8-sig', sep=';')
            print(f"Convertido {excel_file} a {csv_file}")
        except Exception as e:
            print(f"Error al convertir {excel_file} a CSV: {e}")

# Función para obtener el nombre del cliente a partir del nombre del archivo
def obtener_nombre_cliente(filename):
    base = os.path.basename(filename)
    nombre_cliente = base.split('-')[1].strip()
    return nombre_cliente

# Función para obtener la cantidad de faltas de presentación a partir del nombre del archivo
def obtener_faltas_presentacion(filename):
    base = os.path.basename(filename)
    faltas_presentacion = int(base.split('-')[2].strip())
    return faltas_presentacion

# Convertir archivos Excel a CSV
excel_a_csv(input_folder_excel, output_folder_csv)

# Lista para almacenar los DataFrames
df_list = []

# Recorre todos los archivos CSV en la carpeta
for csv_file in glob.glob(os.path.join(output_folder_csv, "*.csv")):
    print(f"Procesando archivo: {csv_file}")
    try:
        # Leer el archivo CSV con la codificación correcta y manejar el BOM
        with open(csv_file, 'r', encoding='utf-8-sig') as file:
            df = pd.read_csv(file, delimiter=';', encoding='latin1', on_bad_lines='skip')
        
        # Eliminar el BOM de los nombres de las columnas
        df.columns = df.columns.str.replace('\ufeff', '')

        # Mostrar las columnas del archivo
        print(f"Columnas del archivo: {df.columns.tolist()}")

        # Verificar si la columna 'Fecha de Vencimiento' existe
        if 'Fecha de Vencimiento' not in df.columns:
            print(f"La columna 'Fecha de Vencimiento' no se encuentra en el archivo: {csv_file}")
            continue  # Saltar este archivo y continuar con el siguiente
        
        # Filtrar por la fecha específica
        df['Fecha de Vencimiento'] = pd.to_datetime(df['Fecha de Vencimiento'], format='%d/%m/%Y', errors='coerce').dt.date
        df = df[df['Fecha de Vencimiento'] < pd.to_datetime(fecha_especifica).date()]
        print(f"Filas después de filtrar por fecha: {len(df)}")

        if df.empty:
            print(f"El archivo {csv_file} no tiene filas después de filtrar por la fecha.")
            continue  # Saltar este archivo si no hay filas después de filtrar
        
        # Convertir columnas a formato numérico
        df['Saldo'] = pd.to_numeric(df['Saldo'].str.replace('.', '').str.replace(',', '.'), errors='coerce')
        df['Int. resarcitorios'] = pd.to_numeric(df['Int. resarcitorios'].str.replace('.', '').str.replace(',', '.'), errors='coerce')
        df['Int. punitorios'] = pd.to_numeric(df['Int. punitorios'].str.replace('.', '').str.replace(',', '.'), errors='coerce')
        
        # Añadir la columna 'Nombre del cliente'
        df['Nombre del cliente'] = obtener_nombre_cliente(csv_file)

        # Añadir la columna 'Cantidad de faltas de presentación'
        df['Cantidad de faltas de presentación'] = obtener_faltas_presentacion(csv_file)

        # Añadir el DataFrame a la lista
        df_list.append(df)

    except Exception as e:
        print(f"Error al procesar el archivo {csv_file}: {e}")

# Concatenar todos los DataFrames en uno solo
if df_list:
    df_resumen = pd.concat(df_list, ignore_index=True)
    
    # Calcular el total de deuda por cliente
    df_totales = df_resumen.groupby('Nombre del cliente')[['Saldo', 'Int. resarcitorios', 'Int. punitorios']].sum().reset_index()
    df_totales['Total deuda'] = df_totales[['Saldo', 'Int. resarcitorios', 'Int. punitorios']].sum(axis=1)

    # Unir los totales con el resumen
    df_resumen = df_resumen.merge(df_totales[['Nombre del cliente', 'Total deuda']], on='Nombre del cliente', how='left')

    # Reordenar columnas
    columnas_ordenadas = ['Nombre del cliente', 'Impuesto', 'Concepto / Subconcepto', 'Ant. / Cuota', 'Período Fiscal', 'Fecha de Vencimiento', 'Saldo', 'Int. resarcitorios', 'Int. punitorios', 'Total deuda', 'Cantidad de faltas de presentación']
    df_resumen = df_resumen[columnas_ordenadas]
    
    # Guardar el DataFrame final en un nuevo archivo CSV
    df_resumen.to_csv(output_file_csv, index=False, encoding='utf-8-sig')

    # Guardar el DataFrame final en un nuevo archivo XLSX
    df_resumen.to_excel(output_file_xlsx, index=False, engine='openpyxl')

    # Lógica para cargar el archivo XLSX y aplicar el formateo
    wb = load_workbook(output_file_xlsx)
    ws = wb.active

    # Definir estilo de borde
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Aplicar estilos a todas las celdas
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
            if cell.column in [7, 8, 9, 10]:  # Columnas de valores numéricos
                cell.number_format = '#,##0.00'

    # Determinar la última fila con datos en cualquier columna
    last_row = ws.max_row 
    for row in reversed(range(1, ws.max_row + 1)):
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=row, column=col).value is not None:
                last_row = row
                break
        if last_row != ws.max_row:
            break
            
    last_row = last_row + 100

    # Combinar celdas y aplicar estilos adicionales
    current_cliente = None
    start_row = 2

    for row in range(2, last_row + 1):
        if ws[f'A{row}'].value != current_cliente:
            if current_cliente is not None:
                end_row = row - 1
                if start_row <= end_row:
                    # Combinar celdas para 'Nombre del cliente'
                    ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
                    # Combinar celdas para 'Total deuda'
                    ws.merge_cells(start_row=start_row, start_column=10, end_row=end_row, end_column=10)
                    # Combinar celdas para 'Cantidad de faltas de presentación'
                    ws.merge_cells(start_row=start_row, start_column=11, end_row=end_row, end_column=11)
                    ws[f'A{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws[f'J{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws[f'K{start_row}'].alignment = Alignment(horizontal='center', vertical='center')

                    # Insertar fila en blanco y colorear en lila claro
                    ws.insert_rows(end_row + 1)
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=end_row + 1, column=col).fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')

            current_cliente = ws[f'A{row}'].value
            start_row = row

    # Manejar la última sección de filas del último cliente
    if current_cliente is not None:
        end_row = last_row
        if start_row <= end_row:
            # Combinar celdas para 'Nombre del cliente'
            ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
            # Combinar celdas para 'Total deuda'
            ws.merge_cells(start_row=start_row, start_column=10, end_row=end_row, end_column=10)
            # Combinar celdas para 'Cantidad de faltas de presentación'
            ws.merge_cells(start_row=start_row, start_column=11, end_row=end_row, end_column=11)
            ws[f'A{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'J{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'K{start_row}'].alignment = Alignment(horizontal='center', vertical='center')

    # Ajustar el ancho de las columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Obtener la letra de la columna
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    wb.save(output_file_xlsx)
    print(f"Resumen de deudas generado en {output_file_xlsx}")

    # Eliminar el archivo CSV
    os.remove(output_file_csv)
else:
    print("No se encontraron archivos con la columna 'Fecha de Vencimiento'.")